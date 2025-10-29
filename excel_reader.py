import openpyxl
from os import listdir
import requests
import re
from os.path import isfile, join
import os

relevant_keywords_total = {"green transition", "green economy", "twin transition", "green jobs", "digitalisation", "green skills","labour market",  "skill needs", "skills forecast",  "digital skills", "labour market information", "skills", "skill needs"}

relevant_keywords_transition = {"green transition", "green economy", "twin transition", "green jobs", "digitalisation", "digital", "digital transition" }

relevant_keywords_skills ={"green skills","labour market",  "skill needs", "skills forecast", "skills", "digital skills", "labour market information", "skill needs", "market", "labour", "skill", "skills forecast"}

folder_to_save= "C:/Users/Evangelia/Documents/ΕΥΑΓΓΕΛΙΑ ΕΓΓΡΑΦΑ/UoCrete/tasks/task 5.1 files/europa/policy_files/"

existing_files = [f for f in listdir(folder_to_save) if isfile(join(folder_to_save, f))]

working_directory= "C:/Users/Evangelia/Documents/ΕΥΑΓΓΕΛΙΑ ΕΓΓΡΑΦΑ/UoCrete/tasks/task 5.1 files/europa/prompts/green transition/2025/"

counter = 0

for f in os.listdir(working_directory):
    print(f)
    if f.endswith(".xlsx"): 

        file_to_read = f 

        # To open the workbook
        # workbook object is created
        wb_obj = openpyxl.load_workbook(os.path.join(working_directory, file_to_read))


        # Get workbook active sheet object
        # from the active attribute
        sheet_obj = wb_obj.active


        row = sheet_obj.max_row
        column = sheet_obj.max_column

        print("Total Rows:", row)
        print("Total Columns:", column)

        descriptions =[]
        titles= []
        urls =[]
        urls_to_files =[]

        # print("\nThese are the descriptions")
        for i in range(2, row + 1):
            cell_obj = sheet_obj.cell(row=i, column=3)
            descriptions.append(cell_obj.value)

        # print("\nThese are the titles")
        for i in range(2, row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)
            titles.append(cell_obj.value)

        # print("\nThese are the urls of the documents")
        for i in range(2, row + 1):
            cell_obj = sheet_obj.cell(row=i, column=8)
            urls.append(cell_obj.value)

        # print("\nThese are the urls that lead to the files")
        for i in range(2, row + 1):
            cell_obj = sheet_obj.cell(row=i, column=9)
            urls_to_files.append(cell_obj.value)

        for i in range(0,len(descriptions)):
            if descriptions[i]:
                file_desc = descriptions[i]
            else:
                file_desc = ""
            
            if titles[i]: 
                file_title = titles[i]
            else: 
                file_title = titles[i]

            if file_title in existing_files:
                print("we skip this one")
                continue
            
            if urls_to_files[i]:
                file_url = urls_to_files[i]
            else:
                file_url=""
            # print(i, descriptions[i], "\n")

            for k in relevant_keywords_skills:
                for l in relevant_keywords_transition:
                    if k in file_desc and l in file_desc:
                        print("yes", k, l)
                        print("Downloading file now...", file_title, "\n")
                        
                        safe_title = re.sub(r'[\\/*?:"<>|]', "_", file_title)
                        safe_title = safe_title.strip().replace("\n", " ")[:150]

                        file_name = os.path.join(folder_to_save, safe_title + ".pdf")

                        if os.path.exists(file_name):
                            print("File already exists, skipping:", file_name)
                            break
                        
                        try:
                            headers = {
                                "User-Agent": "Mozilla/5.0",
                                "Accept": "application/pdf"
                            }
                            r = requests.get(file_url, headers=headers, allow_redirects=True)
                            r.raise_for_status()

                            with open(file_name, "wb") as f:
                                f.write(r.content)

                            print("Downloaded:", file_name, "\n")
                            counter +=1
                        except Exception as e:
                            print("Error downloading", file_url, e, "\n")     
                else: 
                    continue
                break

        print(counter, " new files downloaded")

